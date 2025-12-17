```python
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mega_da_virada.py
Motor do gerador de jogos da Mega-Sena com CLI e funções reutilizáveis.
Uso:
  python mega_da_virada.py 50 --pdf
  from mega_da_virada import gerar_jogos, salva_pdf, carrega_concursos
"""
import random, zipfile, io, json, os, datetime, csv, argparse, sys, time, socket
from urllib.request import urlopen
from urllib.error import URLError

from pathlib import Path

URL_HIST = "https://www1.caixa.gov.br/loterias/_arquivos/loterias/D_megase.zip"
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
CACHE = str(DATA_DIR / "mega_cache.json")

import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def baixa_hist():
    logging.info("Baixando histórico da Caixa...")
    last_exc = None
    for attempt in range(1, 4):
        try:
            data = urlopen(URL_HIST, timeout=20).read()
            z = zipfile.ZipFile(io.BytesIO(data))
            csv_nome = [n for n in z.namelist() if n.upper().endswith('.CSV')][0]
            rows = [ln.decode('ISO-8859-1').strip().split(';') for ln in z.open(csv_nome).readlines()]
            concursos = []
            for r in rows[1:]:
                if len(r) >= 8:
                    try:
                        concursos.append(sorted(map(int, r[2:8])))
                    except ValueError:
                        continue
            with open(CACHE, 'w', encoding='utf8') as f:
                json.dump(concursos, f)
            logging.info(f"Histórico salvo ({len(concursos)} concursos).")
            return concursos
        except (URLError, socket.gaierror) as e:
            last_exc = e
            logging.warning(f"Tentativa {attempt}/3 falhou: {e}")
        except Exception as e:
            last_exc = e
            logging.error(f"Erro processando histórico na tentativa {attempt}: {e}")
        time.sleep(1 * attempt)
    # falhou nas tentativas
    print('Erro ao baixar histórico da Caixa após várias tentativas.')
    print('Motivo:', last_exc)
    print('Verifique: conexão à internet, configuração de proxy/DNS ou bloqueios de firewall.')
    print('Alternativas: importe manualmente o CSV/XLSX oficial e use `carregar_arquivo_local(path)` ou execute `process_local_file.py <arquivo>`.')
    return []

def carrega_concursos():
    if os.path.isfile(CACHE):
        try:
            with open(CACHE, encoding='utf8') as f:
                return json.load(f)
        except Exception:
            return baixa_hist()
    return baixa_hist()

def frequencia(concursos):
    freq = {d:0 for d in range(1,61)}
    for c in concursos:
        for d in c:
            freq[d] += 1
    return freq


def pontuar_dezenas(concursos):
    """Calcula um score simples por dezena baseado em frequência histórica.
    Retorna dict {dezena: score} com score normalizado entre 0 e 1.
    """
    freq = frequencia(concursos)
    maxf = max(freq.values()) if freq else 1
    scores = {d: (freq[d] / maxf) for d in range(1, 61)}
    return scores


def top_dezenas(n=10):
    """Retorna lista dos N pares (dezena, score) ordenados por score desc.
    Lê os concursos do cache atual (se houver).
    """
    concursos = carrega_concursos()
    scores = pontuar_dezenas(concursos)
    ordered = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
    return ordered[:n]


def combined_scores(concursos, recent_n=100, alpha=0.6):
    """Combina frequência histórica e indicador de recência.
    - `alpha` pondera frequência (0..1); recência recebe 1-alpha.
    - `recent_n` define quantos concursos recentes considerar para o componente de recência.
    - Para recência suportamos `decay` linear (padrão) ou `exp` (exponencial).
    Retorna dict {dezena: score} normalizado entre 0 e 1.
    """
    # frequência geral (normalizada)
    freq = frequencia(concursos)
    maxf = max(freq.values()) if freq else 1
    freq_norm = {d: (freq[d] / maxf) for d in range(1, 61)}

    # recência (por padrão linear, mas pode ser ajustada via decay/decay_lambda)
    # NOTE: esta implementação espera parâmetros extras via closure (set por chamadas que incluam decay args).
    # Para compatibilidade, tratamos aqui como linear com pesos por posição.
    recent_list = concursos[-recent_n:] if concursos else []
    rec_raw = {d: 0.0 for d in range(1, 61)}
    if recent_list:
        L = len(recent_list)
        for idx, concurso in enumerate(recent_list):
            # idx: 0..L-1 (mais antigo -> mais novo)
            weight = idx + 1
            for d in concurso:
                rec_raw[d] += weight
    max_raw = max(rec_raw.values()) if rec_raw else 1
    rec_weight = {d: (rec_raw[d] / max_raw) for d in range(1, 61)}

    scores = {}
    for d in range(1, 61):
        scores[d] = alpha * freq_norm.get(d, 0.0) + (1.0 - alpha) * rec_weight.get(d, 0.0)

    # normalize to 0..1
    mx = max(scores.values()) if scores else 1
    if mx <= 0:
        return {d: 0.0 for d in range(1, 61)}
    for d in scores:
        scores[d] = scores[d] / mx
    return scores


def top_dezenas(n=10, use_combined=True):
    """Retorna lista dos N pares (dezena, score) ordenados por score desc.
    Se `use_combined` True, usa `combined_scores` (frequência + recência).
    """
    concursos = carrega_concursos()
    if use_combined:
        try:
            scores = combined_scores(concursos)
        except Exception:
            scores = pontuar_dezenas(concursos)
    else:
        scores = pontuar_dezenas(concursos)
    ordered = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
    return ordered[:n]


def top_dezenas_params(n=10, recent_n=100, alpha=0.6, decay='linear', decay_lambda=0.05):
    """Retorna top N dezenas usando parâmetros: recent_n, alpha, decay e decay_lambda.
    - decay: 'linear' ou 'exp'
    - decay_lambda: taxa para exponencial
    """
    concursos = carrega_concursos()
    try:
        if decay == 'exp':
            # construir recency weights exponenciais manualmente
            recent_list = concursos[-recent_n:] if concursos else []
            raw = {d: 0.0 for d in range(1, 61)}
            if recent_list:
                L = len(recent_list)
                for idx, concurso in enumerate(recent_list):
                    # pos_from_end: 0 = newest
                    pos_from_end = L - 1 - idx
                    weight = pow(2.718281828459045, -decay_lambda * pos_from_end)
                    for d in concurso:
                        raw[d] += weight
            max_raw = max(raw.values()) if raw else 1
            rec_weight = {d: (raw[d] / max_raw) for d in range(1, 61)}
            freq = frequencia(concursos)
            maxf = max(freq.values()) if freq else 1
            freq_norm = {d: (freq[d] / maxf) for d in range(1, 61)}
            scores = {d: alpha * freq_norm.get(d, 0.0) + (1.0 - alpha) * rec_weight.get(d, 0.0) for d in range(1,61)}
            # normalize
            mx = max(scores.values()) if scores else 1
            if mx <= 0:
                scores = {d: 0.0 for d in range(1,61)}
            else:
                for d in scores:
                    scores[d] = scores[d] / mx
        else:
            scores = combined_scores(concursos, recent_n=recent_n, alpha=alpha)
    except Exception:
        scores = pontuar_dezenas(concursos)
    ordered = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
    return ordered[:n]


def frequencies(list_of_lists, max_num=60):
    """Compatibilidade: retorna lista de frequências indexada por dezena (0..max_num)."""
    freq = [0] * (max_num + 1)
    for row in list_of_lists:
        for n in row:
            if 1 <= n <= max_num:
                freq[n] += 1
    return freq

def pesos_invertidos(concursos):
    freq = frequencia(concursos)
    total = sum(freq.values()) or 1
    return {d: (total - f)/total for d,f in freq.items()}

def filtros_ok(jogo, concursos):
    j = sorted(jogo)
    if sum(1 for d in j if d % 2 == 0) >= 4: return False
    for a,b,c in zip(j, j[1:], j[2:]):
        if b == a+1 and c == b+1: return False
    finais = [d % 10 for d in j]
    if max(finais.count(f) for f in set(finais)) > 2: return False
    if len(set(d//10 for d in j)) < 4: return False
    if not (100 <= sum(j) <= 250): return False
    primos = {2,3,5,7,11,13,17,19,23,29,31,37,41,43,47,53,59}
    if sum(1 for d in j if d in primos) > 4: return False
    if j in concursos: return False
    return True

def gerar_jogos(quantidade=20, forcar_filtros=False):
    concursos = carrega_concursos()
    pesos = pesos_invertidos(concursos)
    jogos = []
    while len(jogos) < quantidade:
        if concursos:
            j = sorted(random.choices(range(1,61), weights=[pesos[d] for d in range(1,61)], k=6))
        else:
            j = sorted(random.sample(range(1,61), 6))
        if forcar_filtros or filtros_ok(j, concursos):
            jogos.append(j)
    return jogos

def recomendar_numeros(qtd=6, seed=None, forcar_filtros=False):
    """Gera uma recomendação única de `qtd` dezenas usando os pesos do histórico.
    Retorna lista de inteiros ordenada."""
    if seed is not None:
        random.seed(seed)
    concursos = carrega_concursos()
    pesos = pesos_invertidos(concursos)
    if concursos:
        jogo = sorted(random.choices(range(1,61), weights=[pesos[d] for d in range(1,61)], k=qtd))
    else:
        jogo = sorted(random.sample(range(1,61), qtd))
    if forcar_filtros or filtros_ok(jogo, concursos):
        return jogo
    # tentar novamente algumas vezes com fallback
    for _ in range(1000):
        if concursos:
            jogo = sorted(random.choices(range(1,61), weights=[pesos[d] for d in range(1,61)], k=qtd))
        else:
            jogo = sorted(random.sample(range(1,61), qtd))
        if forcar_filtros or filtros_ok(jogo, concursos):
            return jogo
    return jogo

def custo_aposta(qtd):
    """Retorna o custo em reais (int) de uma aposta com qtd dezenas.
    Fórmula: custo = 6 * C(qtd, 6) quando qtd >= 6, else 0."""
    from math import comb
    if qtd < 6 or qtd > 20:
        return 0
    return 6 * comb(qtd, 6)

def carregar_csv_local(path):
    """Carrega CSV local da Caixa (ou compatível) e atualiza cache JSON.
    Retorna lista de concursos (listas de 6 dezenas)."""
    try:
        with open(path, encoding='utf8') as f:
            rows = [ln.strip().split(';') for ln in f if ln.strip()]
        concursos = []
        for r in rows[1:]:
            try:
                nums = list(map(int, r[1:7]))
                if len(nums) == 6:
                    concursos.append(sorted(nums))
            except Exception:
                continue
        with open(CACHE, 'w', encoding='utf8') as f:
            json.dump(concursos, f)
        return concursos
    except Exception as e:
        raise

def carregar_arquivo_local(path):
    """Carrega CSV ou XLSX local e atualiza o cache.
    Suporta arquivos .csv (ponto-e-vírgula) e .xlsx (openpyxl).
    Retorna lista de concursos (listas de 6 dezenas).
    """
    path = os.path.abspath(path)
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    lower = path.lower()
    concursos = []

    # helper: valida número de dezena
    def is_dezena(v):
        try:
            iv = int(v)
            return 1 <= iv <= 60
        except Exception:
            return False

    # read into rows (list of lists) with header in first row
    rows = []
    header = None
    if lower.endswith('.csv'):
        with open(path, encoding='utf8', errors='ignore') as f:
            for i, ln in enumerate(f):
                parts = [p.strip() for p in ln.strip().split(';')]
                if i == 0:
                    header = parts
                rows.append(parts)
    elif lower.endswith('.xlsx') or lower.endswith('.xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                parts = [str(c).strip() if c is not None else '' for c in row]
                if i == 0:
                    header = parts
                rows.append(parts)
        except Exception:
            try:
                import pandas as pd
                df = pd.read_excel(path)
                header = [str(c) for c in df.columns]
                for _, r in df.iterrows():
                    rows.append([str(r[c]).strip() if not (r[c] is None) else '' for c in df.columns])
            except Exception:
                raise RuntimeError('Instale openpyxl ou pandas para ler .xlsx')
    else:
        raise RuntimeError('Formato não suportado: use .csv ou .xlsx')

    if not rows:
        with open(CACHE, 'w', encoding='utf8') as f:
            json.dump([], f)
        return []

    cols = max(len(r) for r in rows)

    # 1) tentar detectar por cabeçalho comum (D1..D6, D01..D06, DEZENA1..)
    col_candidates = []
    if header:
        lower_hdr = [str(h).strip().lower() for h in header]
        for pattern_idx in range(len(lower_hdr)):
            # try to find sequence of 6 header names matching d1..d6
            names = lower_hdr[pattern_idx:pattern_idx+6]
            if len(names) < 6:
                break
            ok = True
            for j, nm in enumerate(names, start=1):
                if not (nm.startswith(f'd{j}') or nm.startswith(f'd{j:02d}') or nm.startswith(f'dezena') or nm.startswith(f'dez')) and not any(sub in nm for sub in [f'd{j}', f'v{j}', f'{j}']):
                    ok = False
                    break
            if ok:
                col_candidates = list(range(pattern_idx, pattern_idx+6))
                break

    # 2) se não encontrou por cabeçalho, analisar colunas por frequência de inteiros em 1..60
    if not col_candidates:
        counts = [0]*cols
        total_rows = max(1, len(rows)-1)
        # skip header row when counting if header exists
        start_idx = 1 if header else 0
        for r in rows[start_idx:]:
            for ci in range(cols):
                v = r[ci] if ci < len(r) else ''
                if is_dezena(v):
                    counts[ci] += 1
        # pick columns with highest counts
        indexed = sorted(enumerate(counts), key=lambda x: (-x[1], x[0]))
        # select top 6 columns that have at least some numeric hits
        selected = [i for i,c in indexed if c > 0][:6]
        selected.sort()
        if len(selected) == 6:
            col_candidates = selected
        else:
            # fallback: try consecutive window with many numeric entries
            for start in range(0, max(0, cols-6)+1):
                window = list(range(start, start+6))
                score = sum(counts[i] for i in window)
                if score >= (total_rows * 0.1):
                    col_candidates = window
                    break

    # 3) detectar colunas de concurso e data (se existirem)
    concurso_col = None
    data_col = None
    if header:
        lower_hdr = [str(h).strip().lower() for h in header]
        for i, h in enumerate(lower_hdr):
            if any(k in h for k in ('concurso', 'nr', 'numero', 'n_conc', 'num_conc', 'concurso_num')) and concurso_col is None:
                concurso_col = i
            if any(k in h for k in ('data', 'dt', 'sorteio')) and data_col is None:
                data_col = i

    # heurística por conteúdo se não detectado
    def looks_like_date(s):
        if not s:
            return False
        s = s.strip()
        # formatos comuns: dd/mm/yyyy, yyyy-mm-dd, dd-mm-yyyy
        import re
        if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', s):
            return True
        if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', s):
            return True
        return False

    if data_col is None:
        # buscar coluna com muitas ocorrências de datas
        cols = max(len(r) for r in rows)
        date_counts = [0]*cols
        start_idx = 1 if header else 0
        for r in rows[start_idx:]:
            for ci in range(cols):
                v = r[ci] if ci < len(r) else ''
                if looks_like_date(v):
                    date_counts[ci] += 1
        best = sorted(enumerate(date_counts), key=lambda x: -x[1])
        if best and best[0][1] > 0:
            data_col = best[0][0]

    if concurso_col is None:
        # buscar coluna com muitos inteiros crescentes
        cols = max(len(r) for r in rows)
        num_counts = [0]*cols
        start_idx = 1 if header else 0
        for r in rows[start_idx:]:
            for ci in range(cols):
                v = r[ci] if ci < len(r) else ''
                try:
                    iv = int(str(v).strip())
                    if iv > 0:
                        num_counts[ci] += 1
                except Exception:
                    pass
        best = sorted(enumerate(num_counts), key=lambda x: -x[1])
        if best and best[0][1] > 0:
            concurso_col = best[0][0]

    # 4) construir concursos a partir das col_candidates, coletando meta (concurso, data)
    entries = []  # cada item: dict{'concurso': int or None, 'data': str or None, 'dezenas': [6 ints]}
    start_idx = 1 if header else 0
    for r in rows[start_idx:]:
        vals = []
        for ci in col_candidates:
            if ci < len(r) and is_dezena(r[ci]):
                vals.append(int(r[ci]))
            else:
                vals = []
                break
        if len(vals) == 6:
            concurso_val = None
            data_val = None
            if concurso_col is not None and concurso_col < len(r):
                try:
                    concurso_val = int(str(r[concurso_col]).strip())
                except Exception:
                    concurso_val = None
            if data_col is not None and data_col < len(r):
                data_val = str(r[data_col]).strip()
            entries.append({'concurso': concurso_val, 'data': data_val, 'dezenas': sorted(vals)})

    # ordenar por concurso se disponível, senão manter ordem
    try:
        entries_sorted = sorted(entries, key=lambda e: (e['concurso'] if e['concurso'] is not None else 0))
    except Exception:
        entries_sorted = entries

    concursos = [e['dezenas'] for e in entries_sorted]

    # salvar cache padrão
    with open(CACHE, 'w', encoding='utf8') as f:
        json.dump(concursos, f)

    # salvar CSV completo padronizado
    full_path = 'mega_full_from_local.csv'
    with open(full_path, 'w', encoding='utf8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        header_row = ['Concurso', 'Data', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6']
        w.writerow(header_row)
        for e in entries_sorted:
            row = [e['concurso'] if e['concurso'] is not None else '', e['data'] if e['data'] else ''] + e['dezenas']
            w.writerow(row)

    # frequências gerais
    freq = frequencies(concursos, max_num=60)
    freq_path = 'mega_freq_from_local.csv'
    with open(freq_path, 'w', encoding='utf8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['dezena', 'freq'])
        for d in range(1, 61):
            w.writerow([d, freq[d]])

    # frequências recentes (últimos 100 concursos)
    recents = 100
    recent_list = concursos[-recents:] if len(concursos) >= 1 else concursos
    freq_recent = frequencies(recent_list, max_num=60)
    recent_path = 'mega_freq_recent.csv'
    with open(recent_path, 'w', encoding='utf8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['dezena', 'freq_recent'])
        for d in range(1, 61):
            w.writerow([d, freq_recent[d]])

    # resumo JSON
    # calcular top10 usando pontuação combinada (frequência + recência)
    try:
        combined = combined_scores(concursos, recent_n=100, alpha=0.6)
        top10 = sorted(combined.items(), key=lambda x: (-x[1], x[0]))[:10]
    except Exception:
        top10 = sorted([(d, freq[d]) for d in range(1, 61)], key=lambda x: (-x[1], x[0]))[:10]

    summary = {
        'total_concursos': len(concursos),
        'ultima_data': entries_sorted[-1]['data'] if entries_sorted and entries_sorted[-1]['data'] else None,
        'ultima_concurso': entries_sorted[-1]['concurso'] if entries_sorted and entries_sorted[-1]['concurso'] else None,
        'top10': top10
    }
    with open('mega_summary.json', 'w', encoding='utf8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return concursos

def salva_pdf(jogos, arquivo='volantes_mega.pdf'):
    try:
        from fpdf import FPDF
    except Exception:
        print('Biblioteca fpdf não encontrada. Instalando...')
        os.system(f"{sys.executable} -m pip install fpdf")
        from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=10)
    per_page = 12
    for pag in range(0, len(jogos), per_page):
        pdf.add_page()
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(0, 10, f'Mega – Volantes gerados em {datetime.date.today():%d/%m/%Y}', ln=True, align='C')
        pdf.ln(6)
        pdf.set_font('Helvetica', '', 12)
        for i, j in enumerate(jogos[pag:pag+per_page], pag+1):
            pdf.cell(0, 6, f"{i:02d}: " + ' - '.join(f"{d:02d}" for d in j), ln=True)
    pdf.output(arquivo)
    print(f'PDF salvo: {arquivo}')

def salva_csv(jogos, arquivo='volantes_mega.csv'):
    """Salva jogos em CSV simples: Numero;D1;D2;D3;D4;D5;D6"""
    with open(arquivo, 'w', encoding='utf8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Numero','D1','D2','D3','D4','D5','D6'])
        for i, j in enumerate(jogos, 1):
            w.writerow([f'{i:02d}'] + [f'{d:02d}' for d in j])
    print(f'CSV salvo: {arquivo}')

def atualizar_cache():
    """Força download do histórico e atualiza o cache."""
    return baixa_hist()

def main():
    ap = argparse.ArgumentParser(description='Gerador inteligente de jogos da Mega da Virada')
    ap.add_argument('quantidade', type=int, nargs='?', default=20, help='quantos jogos gerar')
    ap.add_argument('--pdf', nargs='?', const='volantes_mega.pdf', help='salva PDF (opcional: nome)')
    ap.add_argument('--csv', nargs='?', const='volantes_mega.csv', help='salva CSV (opcional: nome)')
    ap.add_argument('--forca', action='store_true', help='ignora filtros (gera sem restrições)')
    ap.add_argument('--update', action='store_true', help='força atualização do histórico da Caixa')
    args = ap.parse_args()

    if getattr(args, 'update', False):
        atualizar_cache()

    jogos = gerar_jogos(args.quantidade, forcar_filtros=args.forca)
    for n,j in enumerate(jogos,1):
        print(f"{n:02d}: " + ' - '.join(f"{d:02d}" for d in j))
    if args.pdf:
        salva_pdf(jogos, args.pdf)
    if getattr(args, 'csv', False):
        salva_csv(jogos, args.csv)

if __name__ == '__main__':
    main()

```